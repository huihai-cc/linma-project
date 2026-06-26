import time
import re
import json
import os
import subprocess
import urllib.parse
import openpyxl
from datetime import datetime
from playwright.sync_api import sync_playwright

# ===================================================
# 設定
# ===================================================
USER_DATA_DIR = r"C:\Users\BPO\Desktop\xone_session"

TARGET_URL    = "https://xmediaone.jp/process/dsp/dac-operation/mailBox"
XONE_GAS_URL  = "https://script.google.com/a/macros/dac.co.jp/s/AKfycbwmRUvyZ24xxRWI1_b3rN0f43zYAFfWaRo9aSsmyQJvwLurdLn6DxBcVnSW68r3rOLE/exec"
XONE_PROCESSED_IDS_FILE = r"C:\Users\BPO\Desktop\xone_session\processed_ids.json"
VPN_NAME      = "VPN"

EXCEL_PATH           = r"C:\Users\BPO\Desktop\sharepoint_session\メール転記_5課.xlsx"
MAIL_PROCESSED_IDS_FILE = r"C:\Users\BPO\Desktop\sharepoint_session\processed_ids.json"
SHAREPOINT_SITE_URL  = "https://hakuhodody.sharepoint.com/sites/BPO18"
SHAREPOINT_FILE_URL  = "https://hakuhodody.sharepoint.com/sites/BPO18/Shared%20Documents/%E3%83%A1%E3%83%BC%E3%83%AB%E8%BB%A2%E8%A8%98_5%E8%AA%B2.xlsx"
MAIL_GAS_URL         = "https://script.google.com/a/macros/dac.co.jp/s/AKfycbytUMCDQodxG9ydtcxAzyysvb-MGV6xCjgCppSvjOY9xioWg1HNVwk5T7-ezjr8wAgm/exec"

TEAMS_EXCEL_PATH         = r"C:\Users\BPO\Desktop\xone_session\Teams連絡まとめ_1課.xlsx"
TEAMS_SHAREPOINT_FILE_URL = (
    "https://hakuhodody.sharepoint.com/sites/BPO18"
    "/Shared%20Documents/1%E8%AA%B2"
    "/Teams%E9%80%A3%E7%B5%A1%E3%81%BE%E3%81%A8%E3%82%81_1%E8%AA%B2.xlsx"
    "?download=1"
)
TEAMS_PROCESSED_IDS_FILE = r"C:\Users\BPO\Desktop\xone_session\teams_processed_ids.json"
TEAMS_GAS_URL            = "https://script.google.com/a/macros/dac.co.jp/s/AKfycbwmRUvyZ24xxRWI1_b3rN0f43zYAFfWaRo9aSsmyQJvwLurdLn6DxBcVnSW68r3rOLE/exec"

# ===================================================
# 処理済みID管理
# ===================================================
def load_processed_ids(filepath):
    if os.path.exists(filepath):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                return set(json.load(f))
        except:
            return set()
    return set()

def save_processed_ids(ids, filepath):
    try:
        ids_list = list(ids)
        if len(ids_list) > 500:
            ids_list = ids_list[-500:]
            ids = set(ids_list)

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(ids_list, f, ensure_ascii=False)

        print(f"  💾 ID保存: {len(ids_list)}件")
    except Exception as e:
        print(f"❌ 保存失敗: {e}")

# ===================================================
# GAS送信
# ===================================================
def send_to_gas_xone(gas_page, data):
    try:
        print(f"  GASへ送信中(Xone)...")
        try:
            gas_page.goto("https://script.google.com",
                          wait_until="domcontentloaded",
                          timeout=30000)
            time.sleep(3)
        except Exception as e:
            print(f"  ページ移動: {e}")
            time.sleep(3)

        gas_page.set_default_timeout(120000)

        result = gas_page.evaluate(f"""
        () => {{
            return new Promise((resolve) => {{
                var xhr = new XMLHttpRequest();
                xhr.open("POST", "{XONE_GAS_URL}", true);
                xhr.setRequestHeader("Content-Type", "application/json");
                xhr.timeout = 110000;
                xhr.onload = function() {{
                    resolve("status:" + xhr.status + " body:" + xhr.responseText.substring(0, 300));
                }};
                xhr.onerror = function() {{ resolve("XHRエラー"); }};
                xhr.ontimeout = function() {{ resolve("XHRタイムアウト"); }};
                xhr.send(JSON.stringify({json.dumps(data, ensure_ascii=False)}));
            }});
        }}
        """)

        gas_page.set_default_timeout(30000)
        print(f"  GAS返答(Xone): {result}")
        return True

    except Exception as e:
        print(f"  ❌ GAS送信失敗(Xone): {e}")
        gas_page.set_default_timeout(30000)
        return False


def send_to_gas_mail(gas_page, data):
    try:
        print(f"  GASへ送信中(Mail)...")

        if "body" in data:
            data = data.copy()
            body_str = str(data["body"])[:200]
            # ★追加：=で始まる場合は先頭にスペースを付ける
            if body_str.startswith("="):
                body_str = " " + body_str
            data["body"] = body_str

        data_encoded = urllib.parse.quote(
            json.dumps(data, ensure_ascii=False),
            safe=""
        )
        url = f"{MAIL_GAS_URL}?token=mailsync2024&data={data_encoded}"
        print(f"  URL長: {len(url)}文字")

        gas_page.goto(url, wait_until="domcontentloaded", timeout=60000)
        time.sleep(3)

        body = gas_page.inner_text("body")
        print(f"  GAS返答(Mail): {body[:200]}")

        return "success" in body.lower() or "skip" in body.lower()

    except Exception as e:
        print(f"  ❌ GAS送信失敗(Mail): {e}")
        return False

# ===================================================
# VPN管理
# ===================================================
def is_vpn_connected():
    try:
        result = subprocess.run(['rasdial'], capture_output=True, text=True, timeout=10)
        return VPN_NAME in result.stdout
    except:
        return False

def connect_vpn():
    try:
        if is_vpn_connected():
            print("✅ VPN接続済み")
            return
        print("VPN接続中...")
        subprocess.run(['rasdial', VPN_NAME], capture_output=True, text=True, timeout=30)
        time.sleep(8)
        print("✅ VPN接続完了")
    except Exception as e:
        print(f"VPN接続失敗: {e}")

def disconnect_vpn():
    try:
        if not is_vpn_connected():
            return
        subprocess.run(['rasdial', VPN_NAME, '/disconnect'], capture_output=True, text=True, timeout=10)
        print("✅ VPN切断")
    except Exception as e:
        print(f"VPN切断失敗: {e}")

# ===================================================
# Xone：納期抽出
# ===================================================
def extract_noki(body_text):
    if not body_text:
        return ""
    cut_index = body_text.find("Xone DSP運用案件管理")
    if cut_index != -1:
        body_text = body_text[:cut_index]
    else:
        cut_index2 = body_text.find("転送メッセージ")
        if cut_index2 != -1:
            body_text = body_text[:cut_index2]

    body_text = body_text.replace("（", "(").replace("）", ")")
    body_text = body_text.replace("／", "/")

    patterns = [
        r'\d{1,2}/\d{1,2}(?:\([月火水木金土日]\))?(?:の)?\s*\d{1,2}[:：]\d{2}(?:まで(?:に)?|中)?',
        r'\d{1,2}/\d{1,2}(?:\([月火水木金土日]\))?(?:の)?\s*\d{1,2}時(?:まで(?:に)?|中)?',
        r'\d{1,2}/\d{1,2}(?:\([月火水木金土日]\))?(?:の)?\s*(?:AM|PM|午前|午後)?(?:中|まで(?:に)?)',
        r'(?:本日|明日|今日)(?:の)?\s*(?:\d{1,2}[:：]\d{2}|\d{1,2}時|AM|PM|午前|午後)?(?:中|まで(?:に)?)',
    ]

    results = []
    for pattern in patterns:
        matches = re.findall(pattern, body_text)
        results.extend(matches)

    if results:
        seen = []
        for r in results:
            if r not in seen:
                seen.append(r)
        return "、".join(seen)
    return ""

# ===================================================
# Xone：メール詳細取得
# ===================================================
def get_noki_from_mail(context, page, row):
    new_page = None
    try:
        cells = row.query_selector_all("td")
        if len(cells) < 2:
            return ""

        with context.expect_page(timeout=10000) as new_page_info:
            subject_link = cells[1].query_selector("a.subject")
            if subject_link:
                subject_link.click()
            else:
                cells[1].click()

        new_page = new_page_info.value
        new_page.wait_for_load_state("domcontentloaded", timeout=15000)
        time.sleep(2)

        try:
            new_page.wait_for_selector('p[data-property="bodyHtml"]', state="attached", timeout=10000)
        except:
            return ""

        body_el   = new_page.query_selector('p[data-property="bodyHtml"]')
        body_text = body_el.inner_text() if body_el else ""
        noki      = extract_noki(body_text)

        if noki:
            print(f"  ✅ 納期: {noki}")
        return noki

    except Exception as e:
        print(f"  ⚠️ 読取エラー: {e}")
        return ""
    finally:
        if new_page:
            try:
                new_page.close()
            except:
                pass
        try:
            page.bring_to_front()
        except:
            pass

# ===================================================
# Xone：メインループ
# ===================================================
def run_xone(context, xone_page, gas_page):
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Xone処理開始")

    processed_ids = load_processed_ids(XONE_PROCESSED_IDS_FILE)

    try:
        xone_page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=30000)
    except Exception as e:
        print(f"Xoneナビゲートエラー: {e}")
    time.sleep(3)

    current_url = xone_page.url
    print(f"  現在URL: {current_url}")
    if "mailBox" not in current_url:
        print("⚠️ Xoneログインが必要です")
        input("ログイン後Enterを押してください...")
        time.sleep(3)
        try:
            xone_page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=30000)
        except:
            pass
        time.sleep(3)

    current_url = xone_page.url
    print(f"  遷移後URL: {current_url}")
    if "mailBox" not in current_url:
        print("❌ mailBoxに到達できません。スキップします")
        return

    # 検索実行
    try:
        search_btn = xone_page.get_by_role("button", name="検索", exact=True)
        if search_btn.count() > 0 and search_btn.is_visible():
            search_btn.click()
            print("  検索ボタンクリック")
        else:
            visible_btns = xone_page.locator(
                'button[type="submit"]:not([form="loginForm"])'
            ).filter(has_text=re.compile(r'検索|search', re.IGNORECASE))
            if visible_btns.count() > 0:
                visible_btns.first.click()
                print("  検索ボタン(フィルター)クリック")
            else:
                xone_page.keyboard.press("Enter")
                print("  Enterキーで検索")
    except Exception as e:
        print(f"  検索ボタンエラー: {e}")

    try:
        xone_page.wait_for_selector("tr.am1-table__row", state="attached", timeout=15000)
    except:
        print("⚠️ Xoneテーブル未検出")
        return

    time.sleep(2)

    rows = xone_page.query_selector_all("tr.am1-table__row")
    print(f"  取得行数: {len(rows)}行")
    extracted_data = []

    for row in rows:
        cells = row.query_selector_all("td")
        if len(cells) >= 3:
            date        = cells[0].inner_text().strip()
            subject     = cells[1].inner_text().strip()
            sender_raw  = cells[2].inner_text().strip()
            sender      = sender_raw.split("<")[0].strip() if "<" in sender_raw else sender_raw

            has_digit   = any(char.isdigit() for char in date)
            long_enough = len(subject) > 10
            print(f"  行: date='{date[:15]}' subject='{subject[:20]}' digit={has_digit} len={len(subject)}")

            if not (has_digit and long_enough):
                continue

            mail_id = f"{date}||{subject}"

            if mail_id in processed_ids:
                noki = ""
            else:
                print(f"  メール: {subject[:30]}...")
                noki = get_noki_from_mail(context, xone_page, row)
                processed_ids.add(mail_id)
                save_processed_ids(processed_ids, XONE_PROCESSED_IDS_FILE)

            extracted_data.append([date, subject, sender, noki])

    extracted_data.reverse()
    print(f"  Xone件数: {len(extracted_data)}件")

    if extracted_data:
        connect_vpn()
        send_to_gas_xone(gas_page, {"type": "DSP", "data": extracted_data})

    print(f"[{datetime.now().strftime('%H:%M:%S')}] Xone処理完了")

# ===================================================
# OneDrive：Excelダウンロード
# ===================================================
def download_excel(sp_page):
    try:
        if os.path.exists(EXCEL_PATH):
            os.remove(EXCEL_PATH)

        print("  SharePointからExcelダウンロード中...")
        sp_page.goto(SHAREPOINT_SITE_URL, wait_until="domcontentloaded", timeout=30000)
        time.sleep(2)

        with sp_page.expect_download(timeout=30000) as download_info:
            sp_page.evaluate(f"""
                () => {{
                    const a = document.createElement('a');
                    a.href = '{SHAREPOINT_FILE_URL}';
                    a.download = 'メール転記_5課.xlsx';
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                }}
            """)
        download = download_info.value
        download.save_as(EXCEL_PATH)
        print("  ✅ ダウンロード完了")
        time.sleep(2)
        return True

    except Exception as e:
        print(f"  ❌ ダウンロード失敗: {e}")
        if os.path.exists(EXCEL_PATH):
            print("  既存ファイルを使用します")
            return True
        return False

# ===================================================
# OneDrive：本文キーワード抽出
# ===================================================
def extract_key_body(body):
    if not body:
        return ""

    signature_marker = "株式会社Hakuhodo DY ONE"
    if signature_marker in body:
        body = body[:body.index(signature_marker)]

    lines = body.splitlines()
    keywords = [
        "GP本数", "グループ本数", "セット本数",
        "広告本数", "追加広告本数", "ポスト本数",
        "設定依頼本数",
        "▼設定スケジュール",
        "Sprinklr",
        "設定スケジュール",
        "SNS",
        "LOCALIO_ORDER",
    ]

    result_lines     = []
    capture_schedule = False
    jno_list         = []

    for line in lines:
        # ★追加：=が連続する分隔线をスキップ
        if re.match(r'^[=\-－＝]{3,}$', line.strip()):
            continue

        if "▼設定スケジュール" in line:
            capture_schedule = True
            result_lines.append(line)
            continue

        if capture_schedule:
            if "===" in line:
                capture_schedule = False
            else:
                result_lines.append(line)
            continue

        if re.search(r'jno', line, re.IGNORECASE):
            matches = re.findall(r'\bJ\d{5,}\b', line)
            if matches:
                jno_list.extend(matches)
            continue

        standalone = re.findall(r'\bJ\d{5,}\b', line)
        if standalone and not any(kw in line for kw in keywords):
            jno_list.extend(standalone)
            continue

        for kw in keywords:
            if kw in line:
                result_lines.append(line)
                break

    if jno_list:
        seen = []
        for j in jno_list:
            if j not in seen:
                seen.append(j)
        result_lines.append("JNO: " + "、".join(seen))

    # ★追加：先頭が=で始まる行を安全な文字に置換
    safe_lines = []
    for line in result_lines:
        if line.startswith("="):
            line = "'" + line  # 先頭に'を付けて数式扱いを防ぐ
        safe_lines.append(line)

    return "\n".join(safe_lines)

# ===================================================
# Teams：Excelダウンロード
# ===================================================
def download_teams_excel(sp_page):
    try:
        if os.path.exists(TEAMS_EXCEL_PATH):
            os.remove(TEAMS_EXCEL_PATH)

        sp_page.goto(
            "https://hakuhodody.sharepoint.com/sites/BPO18",
            wait_until="domcontentloaded",
            timeout=30000
        )
        time.sleep(2)

        with sp_page.expect_download(timeout=60000) as download_info:
            sp_page.evaluate(f"""
                () => {{
                    const a = document.createElement('a');
                    a.href = '{TEAMS_SHAREPOINT_FILE_URL}';
                    a.click();
                }}
            """)

        download = download_info.value
        download.save_as(TEAMS_EXCEL_PATH)
        print("  ✅ Teamsファイルダウンロード完了")
        time.sleep(2)
        return True

    except Exception as e:
        print(f"  ❌ Teamsダウンロード失敗: {e}")
        if os.path.exists(TEAMS_EXCEL_PATH):
            return True
        return False

# ===================================================
# Teams：GAS送信
# ===================================================
def send_to_gas_teams(gas_page, data):
    try:
        print(f"  GASへ送信中(Teams)...")

        try:
            gas_page.goto(
                "https://script.google.com",
                wait_until="domcontentloaded",
                timeout=30000
            )
            time.sleep(3)
        except Exception as e:
            print(f"  ページ移動: {e}")
            time.sleep(3)

        gas_page.set_default_timeout(120000)
        payload = json.dumps(data, ensure_ascii=False)

        result = gas_page.evaluate(f"""
        () => {{
            return new Promise((resolve) => {{
                var xhr = new XMLHttpRequest();
                xhr.open("POST", "{TEAMS_GAS_URL}", true);
                xhr.setRequestHeader("Content-Type", "application/json");
                xhr.timeout = 110000;
                xhr.onload = function() {{
                    resolve("status:" + xhr.status + " body:" + xhr.responseText.substring(0, 300));
                }};
                xhr.onerror = function() {{ resolve("XHRエラー"); }};
                xhr.ontimeout = function() {{ resolve("XHRタイムアウト"); }};
                xhr.send(JSON.stringify({payload}));
            }});
        }}
        """)

        gas_page.set_default_timeout(30000)
        print(f"  GAS返答(Teams): {result}")
        return "SUCCESS" in result or "success" in result.lower()

    except Exception as e:
        print(f"  ❌ GAS送信失敗(Teams): {e}")
        gas_page.set_default_timeout(30000)
        return False

# ===================================================
# Teams：メインループ
# ===================================================
def run_teams_transfer(sp_page, gas_page):
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Teams処理開始")

    processed_ids = load_processed_ids(TEAMS_PROCESSED_IDS_FILE)
    print(f"  既処理件数: {len(processed_ids)}件")

    if not download_teams_excel(sp_page):
        print("  ❌ Excel取得失敗、スキップ")
        return

    try:
        wb = openpyxl.load_workbook(TEAMS_EXCEL_PATH)
        print(f"  シート一覧: {wb.sheetnames}")
        ws = wb["フロント3入稿"]
    except Exception as e:
        print(f"  ❌ Excel読込失敗: {e}")
        return

    rows_to_send = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        received_time = row[0]
        team_name     = row[1]
        channel       = row[2]
        sender        = row[3]
        subject       = row[4]
        body          = row[5]
        teams_url     = row[6]

        if not received_time:
            continue

        mail_id = str(received_time).strip()

        if mail_id in processed_ids:
            continue

        print(f"  新規: {mail_id} | {str(subject)[:30]}")

        rows_to_send.append({
            "receivedTime": mail_id,
            "teamName"    : str(team_name).strip() if team_name else "",
            "channel"     : str(channel).strip()   if channel   else "",
            "sender"      : str(sender).strip()    if sender    else "",
            "subject"     : str(subject).strip()   if subject   else "",
            "body"        : str(body).strip()[:50000] if body   else "",
            "teamsUrl"    : str(teams_url).strip() if teams_url else ""
        })

    print(f"  新規件数: {len(rows_to_send)}件")

    if rows_to_send:
        chunk_size   = 50
        failed_items = []

        for i in range(0, len(rows_to_send), chunk_size):
            chunk = rows_to_send[i:i + chunk_size]
            print(f"  送信中: {i+1}〜{i+len(chunk)}件目")
            success = send_to_gas_teams(gas_page, {"type": "TEAMS", "data": chunk})

            if success:
                print(f"  ✅ 送信成功")
                # ✅ 成功した分のIDだけ保存
                for item in chunk:
                    processed_ids.add(item["receivedTime"])
            else:
                print(f"  ❌ 送信失敗 → 次回再試行")
                failed_items.extend(chunk)

            time.sleep(3)

    # ✅ 成功分だけ保存（失敗分は次回再試行）
    save_processed_ids(processed_ids, TEAMS_PROCESSED_IDS_FILE)

    print(f"[{datetime.now().strftime('%H:%M:%S')}] Teams処理完了")

# ===================================================
# OneDrive：メインループ
# ===================================================
def run_mail_transfer(sp_page, gas_page):
    print(f"\n[{datetime.now().strftime('%H:%M:%S')}] OneDrive処理開始")

    processed_ids = load_processed_ids(MAIL_PROCESSED_IDS_FILE)
    print(f"  既処理件数: {len(processed_ids)}件")

    if not download_excel(sp_page):
        print("  ❌ Excel取得失敗")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["メール転記"]
    except Exception as e:
        print(f"  ❌ Excel読込失敗: {e}")
        return

    new_ids    = set()
    sent_count = 0
    row_count  = 0

    for row in ws.iter_rows(min_row=2, values_only=False):
        received_time  = row[0].value
        subject        = row[1].value
        sender_email   = row[2].value
        body           = row[3].value
        status_cell    = row[4]
        current_status = str(status_cell.value).strip() if status_cell.value else ""

        if not subject:
            continue

        row_count  += 1
        subject_str = str(subject).strip()
        mail_id     = f"{received_time}||{subject_str}"

        if mail_id in processed_ids:
            continue

        if current_status and current_status not in ["未処理", "None", ""]:
            new_ids.add(mail_id)
            continue

        print(f"  送信: {subject_str[:50]}")

        row_data = {
            "receivedTime": str(received_time) if received_time else "",
            "subject"     : subject_str,
            "sender"      : str(sender_email).strip() if sender_email else "",
            "body"        : extract_key_body(str(body)) if body else ""
        }

        success = send_to_gas_mail(gas_page, row_data)

        if success:
            new_ids.add(mail_id)
            sent_count += 1
            time.sleep(2)
        else:
            print(f"  ⚠️ 失敗、次回再試行")

    print(f"  Excel総行数: {row_count}件")
    print(f"  送信完了: {sent_count}件")

    if new_ids:
        processed_ids.update(new_ids)
        save_processed_ids(processed_ids, MAIL_PROCESSED_IDS_FILE)

    print(f"[{datetime.now().strftime('%H:%M:%S')}] OneDrive処理完了")

# ===================================================
# 起動
# ===================================================
def run():
    with sync_playwright() as p:
        print("=== 統合Bot 起動 ===")
        context = p.chromium.launch_persistent_context(
            user_data_dir=USER_DATA_DIR,
            headless=False,
            args=['--start-maximized', '--no-first-run'],
            ignore_default_args=['--enable-automation'],
            channel="chrome"
        )

        xone_page = context.new_page()
        sp_page   = context.new_page()
        gas_page  = context.new_page()

        # Xoneログイン確認
        print("Xoneログイン確認中...")
        xone_page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=30000)
        time.sleep(3)
        if "mailBox" not in xone_page.url:
            print("⚠️ Xoneログインが必要です")
            input("ログイン後Enterを押してください...")
            xone_page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=30000)
            time.sleep(3)

        # SharePointログイン確認
        print("SharePointログイン確認中...")
        sp_page.goto(SHAREPOINT_SITE_URL, wait_until="domcontentloaded", timeout=30000)
        time.sleep(3)
        if "login" in sp_page.url or "microsoftonline" in sp_page.url:
            print("⚠️ SharePointログインが必要です")
            input("ログイン後Enterを押してください...")

        # GASページ準備
        print("GASページ準備中...")
        gas_page.goto("https://script.google.com", wait_until="domcontentloaded", timeout=30000)
        time.sleep(3)
        if "accounts.google" in gas_page.url:
            print("⚠️ Googleログインが必要です")
            input("ログイン後Enterを押してください...")

        print("✅ 全ログイン確認完了\n")

        xone_counter  = 0
        teams_counter = 0

        while True:
            try:
                run_mail_transfer(sp_page, gas_page)

                if xone_counter % 5 == 0:
                    run_xone(context, xone_page, gas_page)

                if teams_counter % 2 == 0:
                    run_teams_transfer(sp_page, gas_page)

                xone_counter  += 1
                teams_counter += 1

            except Exception as e:
                print(f"❌ エラー: {e}")

            print("  1分後に再実行...\n")
            time.sleep(60)

if __name__ == "__main__":
    run()